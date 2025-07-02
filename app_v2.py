import streamlit as st
import gspread
import os
from datetime import datetime
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import io
import json
import tempfile
import requests
import concurrent.futures
import threading

class BuscadorPlacasWeb:
    def __init__(self):
        self.gc = None
        self.credenciales_path = None
        if 'resultados_actuales' not in st.session_state:
            st.session_state.resultados_actuales = []
        self.detectar_credenciales()
    
    def detectar_credenciales(self):
        """Detecta credenciales desde secrets o archivo local"""
        # Primero intentar desde Streamlit secrets (para producción en Streamlit Cloud)
        try:
            if hasattr(st, 'secrets') and 'gcp_service_account' in st.secrets:
                # Crear archivo temporal con las credenciales
                credentials = dict(st.secrets['gcp_service_account'])
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                    json.dump(credentials, f)
                    self.credenciales_path = f.name
                return
        except Exception:
            pass
        
        # Fallback a archivo local (para desarrollo)
        archivos_json = glob.glob("*.json")
        if archivos_json:
            self.credenciales_path = archivos_json[0]
    
    def conectar_google_sheets(self):
        """Conecta con Google Sheets"""
        try:
            if not self.credenciales_path or not os.path.exists(self.credenciales_path):
                raise FileNotFoundError("Archivo de conexión no encontrado")
            
            self.gc = gspread.service_account(filename=self.credenciales_path)
            return True
        except Exception as e:
            st.error(f"Error de conexión: {str(e)}")
            return False
    
    def buscar_placas_en_drive(self, placa_buscar):
        """Busca una placa en todas las hojas RRV"""
        if not self.gc:
            if not self.conectar_google_sheets():
                return []
        
        try:
            with st.spinner('Buscando en Google Sheets...'):
                todas_las_hojas = self.gc.openall()
                hojas_rrv = [hoja for hoja in todas_las_hojas if "RRV" in hoja.title]
                
                if not hojas_rrv:
                    st.warning("No se encontraron hojas con 'RRV' en el nombre")
                    return []
                
                resultados = []
                progress_bar = st.progress(0)
                
                for idx, hoja in enumerate(hojas_rrv):
                    try:
                        worksheets = hoja.worksheets()
                        
                        for worksheet in worksheets:
                            try:
                                data = worksheet.get_all_values()
                                if not data or len(data) < 2:
                                    continue
                                
                                encabezados = data[0]
                                filas_datos = data[1:]
                                
                                filas_encontradas = self.buscar_placa_en_hoja(
                                    filas_datos, encabezados, placa_buscar, hoja.title, worksheet.title
                                )
                                
                                if filas_encontradas:
                                    resultados.extend(filas_encontradas)
                                    
                            except Exception as e:
                                continue
                                
                    except Exception as e:
                        continue
                    
                    progress_bar.progress((idx + 1) / len(hojas_rrv))
                
                progress_bar.empty()
                return resultados
                
        except Exception as e:
            st.error(f"Error durante la búsqueda: {str(e)}")
            return []
    
    def buscar_placa_en_hoja(self, filas_datos, encabezados, placa_buscar, nombre_spreadsheet, nombre_worksheet):
        """Busca una placa en una hoja específica"""
        resultados = []
        
        # Buscar columnas de placa
        columnas_placa = []
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['placa', 'patente', 'matricula', 'vehiculo', 'numero de vehiculo']):
                columnas_placa.append(i)
        
        if not columnas_placa:
            columnas_placa = list(range(min(3, len(encabezados))))
        
        # Buscar la placa
        for num_fila, fila in enumerate(filas_datos, start=2):
            for col_placa in columnas_placa:
                if col_placa < len(fila):
                    valor_celda = str(fila[col_placa]).strip()
                    if placa_buscar.upper() in valor_celda.upper():
                        # Encontrar columnas específicas
                        fecha_col = self.encontrar_columna_fecha(encabezados)
                        proyecto_col = self.encontrar_columna_proyecto(encabezados)
                        empresa_col = self.encontrar_columna_empresa(encabezados)
                        sistema_col = self.encontrar_columna_sistema(encabezados)
                        trabajo_col = self.encontrar_columna_trabajo(encabezados)
                        
                        resultado = {
                            'hoja': nombre_spreadsheet,
                            'pestana': nombre_worksheet,
                            'fila': num_fila,
                            'placa': valor_celda,
                            'fecha': fila[fecha_col] if fecha_col < len(fila) else "No disponible",
                            'proyecto': fila[proyecto_col] if proyecto_col < len(fila) else "No disponible",
                            'empresa': fila[empresa_col] if empresa_col < len(fila) else "No disponible",
                            'sistema': fila[sistema_col] if sistema_col < len(fila) else "No disponible",
                            'trabajo': fila[trabajo_col] if trabajo_col < len(fila) else "No disponible",
                            'datos_completos': fila,
                            'encabezados': encabezados
                        }
                        resultados.append(resultado)
                        break
        
        return resultados
    
    def encontrar_columna_fecha(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['fecha', 'date', 'dia', 'hora', 'fecha de ingreso']):
                return i
        return 1 if len(encabezados) > 1 else 0
    
    def encontrar_columna_proyecto(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if 'proyecto' in encabezado_lower:
                return i
        return 2 if len(encabezados) > 2 else 0
    
    def encontrar_columna_empresa(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['empresa', 'nombre', 'cliente']):
                return i
        return 3 if len(encabezados) > 3 else 0
    
    def encontrar_columna_sistema(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if 'sistema' in encabezado_lower:
                return i
        return 4 if len(encabezados) > 4 else 0
    
    def encontrar_columna_trabajo(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['tipo de trabajo', 'estado', 'status', 'situacion', 'condicion']):
                return i
        return 5 if len(encabezados) > 5 else 0
    
    def ordenar_resultados_cronologicamente(self, resultados):
        """Ordena los resultados por fecha de manera cronológica"""
        def parsear_fecha(fecha_str):
            """Intenta parsear diferentes formatos de fecha"""
            if not fecha_str or fecha_str == "No disponible":
                return datetime.min
            
            fecha_str = str(fecha_str).strip()
            
            # Limpiar la fecha de caracteres extra
            fecha_str = fecha_str.replace('  ', ' ').strip()
            
            # Formatos comunes de fecha (ordenados de más específico a más general)
            formatos = [
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%d',
                '%d-%m-%Y %H:%M:%S',
                '%d-%m-%Y %H:%M',
                '%d-%m-%Y',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %H:%M',
                '%m/%d/%Y',
                '%d/%m/%y %H:%M:%S',
                '%d/%m/%y %H:%M',
                '%d/%m/%y',
                '%d.%m.%Y %H:%M:%S',
                '%d.%m.%Y %H:%M',
                '%d.%m.%Y'
            ]
            
            for formato in formatos:
                try:
                    return datetime.strptime(fecha_str, formato)
                except ValueError:
                    continue
            
            # Si no se puede parsear, intentar con dateutil
            try:
                from dateutil import parser
                return parser.parse(fecha_str)
            except:
                pass
            
            # Si no se puede parsear, devolver fecha mínima
            return datetime.min
        
        # Ordenar por fecha (más reciente primero)
        resultados_ordenados = sorted(
            resultados, 
            key=lambda x: parsear_fecha(x['fecha']), 
            reverse=True
        )
        
        return resultados_ordenados
    
    def consultar_api_rrvsac(self, placa):
        """Consulta la API de RRVSAC para verificar el estado de una placa"""
        try:
            url = 'https://plataforma.rrvsac.com/api/vehicles'
            params = {'search.info.license_plate': placa.strip()}
            headers = {
                'authenticate': 'e843453d60c9b826ed4704f77a88ab6fb4bcb9cd88b2ce25e600cd5b',
                'Accept': '*/*',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive'
            }
            response = requests.get(url, params=params, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                # Buscar campo 'id' dentro de la estructura 'data'
                if data and isinstance(data, dict) and 'data' in data:
                    data_content = data['data']
                    # Si data es una lista, buscar en el primer elemento
                    if isinstance(data_content, list) and data_content:
                        first_item = data_content[0]
                        if isinstance(first_item, dict) and 'id' in first_item:
                            return 'ACTIVO'
                    # Si data es un diccionario, buscar directamente
                    elif isinstance(data_content, dict) and 'id' in data_content:
                        return 'ACTIVO'
                return 'NO ACTIVO'
            else:
                return 'NO ACTIVO'
        except Exception as e:
            st.error(f"Error al consultar la API de RRVSAC: {str(e)}")
            return 'NO ACTIVO'
    
    def crear_excel_bytes(self, resultado):
        """Crea un archivo Excel en memoria y devuelve los bytes"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Placa {resultado['placa']}"
            
            # Estilos
            titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            subtitulo_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=10)
            
            titulo_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
            subtitulo_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws['A1'] = f"INFORMACIÓN DE LA PLACA: {resultado['placa']}"
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws.merge_cells('A1:C1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Información de ubicación
            ws['A3'] = "UBICACIÓN DEL REGISTRO"
            ws['A3'].font = subtitulo_font
            ws['A3'].fill = subtitulo_fill
            ws['A3'].border = border
            
            ws['A4'] = "Hoja:"
            ws['B4'] = resultado['hoja']
            ws['A5'] = "Pestaña:"
            ws['B5'] = resultado['pestana']
            ws['A6'] = "Fila:"
            ws['B6'] = resultado['fila']
            
            # Aplicar estilos a la información de ubicación
            for row in range(4, 7):
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].border = border
                ws[f'B{row}'].font = normal_font
                ws[f'B{row}'].border = border
            
            # Datos completos de la fila
            ws['A8'] = "DATOS COMPLETOS DE LA FILA"
            ws['A8'].font = subtitulo_font
            ws['A8'].fill = subtitulo_fill
            ws['A8'].border = border
            ws.merge_cells('A8:C8')
            
            # Encabezados de la tabla de datos
            ws['A9'] = "Campo"
            ws['B9'] = "Valor"
            ws['A9'].font = subtitulo_font
            ws['B9'].font = subtitulo_font
            ws['A9'].fill = subtitulo_fill
            ws['B9'].fill = subtitulo_fill
            ws['A9'].border = border
            ws['B9'].border = border
            
            # Insertar todos los datos de la fila
            row_num = 10
            for encabezado, valor in zip(resultado['encabezados'], resultado['datos_completos']):
                ws[f'A{row_num}'] = encabezado
                ws[f'B{row_num}'] = valor
                ws[f'A{row_num}'].font = normal_font
                ws[f'B{row_num}'].font = normal_font
                ws[f'A{row_num}'].border = border
                ws[f'B{row_num}'].border = border
                row_num += 1
            
            # Información del archivo
            ws[f'A{row_num + 1}'] = f"Archivo generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws[f'A{row_num + 1}'].font = Font(name='Arial', size=9, italic=True)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            wb.close()
            
            return output.getvalue()
        except Exception as e:
            st.error(f"Error al crear archivo Excel: {str(e)}")
            return None

def main():
    # Configuración de la página con tema oscuro forzado
    st.set_page_config(
        page_title="🔍 Buscador RRV",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Forzar tema oscuro desde el inicio
    st.markdown("""
    <script>
    const doc = window.parent.document;
    doc.body.classList.add('dark-theme');
    </script>
    """, unsafe_allow_html=True)
    
    # CSS tema oscuro moderno y elegante
    st.markdown("""
    <style>
    :root {
        --bg-primary: #0f1419;
        --bg-secondary: #1a1f26;
        --bg-tertiary: #252a32;
        --text-primary: #f8fafc;
        --text-secondary: #cbd5e1;
        --text-muted: #64748b;
        --accent-primary: #3b82f6;
        --accent-secondary: #ef4444;
        --accent-success: #10b981;
        --border-color: #374151;
        --shadow-dark: 0 4px 12px rgba(0,0,0,0.4);
        --shadow-light: 0 2px 8px rgba(0,0,0,0.2);
        --gradient-primary: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        --gradient-secondary: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
        --gradient-success: linear-gradient(135deg, #10b981 0%, #059669 100%);
    }
    
    /* Forzar tema oscuro en toda la aplicación */
    html, body, .stApp, .main, .block-container {
        background: var(--bg-primary) !important;
        color: var(--text-primary) !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }
    
    /* Sobrescribir todos los elementos de Streamlit */
    .stApp * {
        color: var(--text-primary) !important;
    }
    
    /* Forzar fondo oscuro en todos los contenedores */
    .stApp > div, .stApp > div > div, .main > div, .block-container > div {
        background: var(--bg-primary) !important;
    }
    
    .main-header {
        background: var(--gradient-primary);
        padding: 2.5rem 2rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
        box-shadow: var(--shadow-dark);
        border: 1px solid var(--border-color);
    }
    
    .main-header h1 {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0 0 0.5rem 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    
    .main-header p {
        font-size: 1.2rem;
        margin: 0;
        opacity: 0.9;
    }
    
    .search-container {
        background: var(--bg-secondary);
        padding: 2.5rem;
        border-radius: 16px;
        box-shadow: var(--shadow-dark);
        margin-bottom: 2rem;
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        backdrop-filter: blur(10px);
    }
    
    .results-container {
        background: var(--bg-secondary);
        padding: 2.5rem;
        border-radius: 16px;
        box-shadow: var(--shadow-dark);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        backdrop-filter: blur(10px);
    }
    
    .stButton > button {
        background: var(--gradient-secondary) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        font-weight: 600 !important;
        padding: 0.875rem 2rem !important;
        font-size: 1rem !important;
        transition: all 0.3s ease !important;
        box-shadow: var(--shadow-light) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: var(--shadow-dark) !important;
    }
    
    .stTextInput > div > div > input {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
        border: 2px solid var(--border-color) !important;
        border-radius: 12px !important;
        padding: 1rem 1.25rem !important;
        font-size: 1rem !important;
        transition: all 0.3s ease !important;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: var(--accent-primary) !important;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15) !important;
    }
    
    .stAlert {
        border-radius: 12px !important;
        border: none !important;
        box-shadow: var(--shadow-light) !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stSuccess {
        background: rgba(16, 185, 129, 0.1) !important;
        color: var(--accent-success) !important;
        border-left: 4px solid var(--accent-success) !important;
    }
    
    .stWarning {
        background: rgba(245, 158, 11, 0.1) !important;
        color: #f59e0b !important;
        border-left: 4px solid #f59e0b !important;
    }
    
    .stError {
        background: rgba(239, 68, 68, 0.1) !important;
        color: var(--accent-secondary) !important;
        border-left: 4px solid var(--accent-secondary) !important;
    }
    
    .stInfo {
        background: rgba(59, 130, 246, 0.1) !important;
        color: var(--accent-primary) !important;
        border-left: 4px solid var(--accent-primary) !important;
    }
    
    .stDataFrame {
        background: var(--bg-tertiary) !important;
        border-radius: 12px !important;
        overflow: hidden !important;
        box-shadow: var(--shadow-light) !important;
        border: 1px solid var(--border-color) !important;
    }
    
    /* Forzar tema oscuro en tablas */
    .stDataFrame table, .stDataFrame thead, .stDataFrame tbody, .stDataFrame tr, .stDataFrame td, .stDataFrame th {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
        border-color: var(--border-color) !important;
    }
    
    .stDataFrame th {
        background: var(--bg-secondary) !important;
        color: var(--text-primary) !important;
        font-weight: 600 !important;
    }
    
    .stExpander {
        background: var(--bg-tertiary) !important;
        border-radius: 12px !important;
        border: 1px solid var(--border-color) !important;
        margin-bottom: 1rem !important;
        box-shadow: var(--shadow-light) !important;
    }
    
    /* Contenido del expander */
    .stExpander > div > div {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
    }
    
    .stExpanderHeader {
        color: var(--accent-primary) !important;
        font-weight: 600 !important;
        padding: 1rem !important;
        border-radius: 12px !important;
    }
    
    .stMetric {
        background: var(--bg-tertiary) !important;
        border-radius: 12px !important;
        padding: 1.5rem !important;
        border: 1px solid var(--border-color) !important;
        box-shadow: var(--shadow-light) !important;
    }
    
    .stMetric > div {
        color: var(--text-primary) !important;
    }
    
    .stMetric [data-testid="metric-value"] {
        color: var(--accent-primary) !important;
        font-size: 2rem !important;
        font-weight: 700 !important;
    }
    
    .stSubheader {
        color: var(--text-primary) !important;
        font-weight: 600 !important;
        margin-bottom: 1rem !important;
    }
    
    /* Forzar colores de texto en todos los elementos */
    p, h1, h2, h3, h4, h5, h6, .stMarkdown, .stText, .stCaption, label, span, div {
        color: var(--text-primary) !important;
    }
    
    /* Elementos específicos de Streamlit */
    .stSelectbox label, .stTextInput label, .stTextArea label, .stNumberInput label {
        color: var(--text-primary) !important;
        font-weight: 500 !important;
    }
    
    /* Forzar fondo en elementos de formulario */
    .stSelectbox > div > div, .stTextArea > div > div > textarea {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
        border: 2px solid var(--border-color) !important;
    }
    
    .stColumns > div {
        padding: 0 0.5rem !important;
    }
    
    /* Scrollbar personalizada */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--bg-secondary);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--accent-primary);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--accent-secondary);
    }
    
    /* Status badges */
    .status-active {
        background: var(--gradient-success) !important;
        color: white !important;
        padding: 8px 24px !important;
        border-radius: 50px !important;
        font-weight: 600 !important;
        font-size: 1.1rem !important;
        box-shadow: var(--shadow-light) !important;
        display: inline-block !important;
    }
    
    .status-inactive {
        background: var(--gradient-secondary) !important;
        color: white !important;
        padding: 8px 24px !important;
        border-radius: 50px !important;
        font-weight: 600 !important;
        font-size: 1.1rem !important;
        box-shadow: var(--shadow-light) !important;
        display: inline-block !important;
    }
    
    /* Footer styling */
    .footer {
        text-align: center;
        color: var(--text-muted);
        font-size: 0.9rem;
        padding: 2rem 0;
        border-top: 1px solid var(--border-color);
        margin-top: 3rem;
    }
    
    /* Animaciones suaves */
    .search-container, .results-container, .main-header {
        transition: all 0.3s ease;
    }
    
    .search-container:hover, .results-container:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.5);
    }
    
    /* Eliminar completamente el tema claro de Streamlit */
    .stApp [data-testid="stSidebar"] {
        background: var(--bg-secondary) !important;
    }
    
    /* Forzar tema oscuro en elementos específicos */
    .stSpinner, .stProgress, .stSlider {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
    }
    
    /* Eliminar fondos blancos residuales */
    .element-container, .stBlock, .stColumn {
        background: transparent !important;
    }
    
    /* Forzar tema oscuro en markdown y código */
    .stMarkdown pre, .stMarkdown code {
        background: var(--bg-secondary) !important;
        color: var(--text-primary) !important;
        border: 1px solid var(--border-color) !important;
    }
    
    /* Elementos de navegación y menú */
    .stTabs, .stTab {
        background: var(--bg-secondary) !important;
        color: var(--text-primary) !important;
    }
    
    /* Spinner y elementos de carga */
    .stSpinner > div {
        border-color: var(--accent-primary) !important;
    }
    
    /* Elementos de fecha y hora */
    .stDateInput, .stTimeInput {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header principal
    st.markdown("""
    <div class="main-header">
        <h1>🔍 BUSCADOR RRV</h1>
        <p>Sistema de búsqueda de placas en Google Sheets</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Inicializar la aplicación
    app = BuscadorPlacasWeb()
    
    # Verificar credenciales
    if not app.credenciales_path:
        st.error("❌ No se encontraron credenciales. Contacta al administrador para configurar el acceso.")
        st.info("💡 Para desarrolladores: Configura las credenciales en Streamlit Cloud Secrets o agrega un archivo JSON local.")
        return
    
    st.success("✅ Sistema conectado y listo para buscar.")
    
    # Buscar Placa
    st.markdown('<div class="search-container">', unsafe_allow_html=True)
    st.subheader("📋 Buscar Placa")
    col1, col2 = st.columns([3, 1])
    with col1:
        placa_buscar = st.text_input(
            "Ingresa la placa a buscar:",
            placeholder="Ej: ABC-123",
            key="placa_input"
        )
    with col2:
        st.write("")  # Espaciado
        buscar_btn = st.button("🔍 Buscar", type="primary", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Ejecutar búsquedas en paralelo
    if buscar_btn and placa_buscar.strip():
        with st.spinner('🔍 Buscando en Google Sheets y consultando API de RRVSAC en paralelo...'):
            # Ejecutar ambas búsquedas en paralelo
            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                # Búsqueda en Google Sheets
                future_sheets = executor.submit(app.buscar_placas_en_drive, placa_buscar.strip())
                # Consulta a la API de RRVSAC
                future_api = executor.submit(app.consultar_api_rrvsac, placa_buscar.strip())
                
                # Obtener resultados
                resultados = future_sheets.result()
                rrvsac_status = future_api.result()
        
        # Procesar resultados de Google Sheets
        resultados_ordenados = app.ordenar_resultados_cronologicamente(resultados)
        st.session_state.resultados_actuales = resultados_ordenados
        
        if not resultados_ordenados:
            st.warning("⚠️ No se encontró esta placa en el sistema")
        else:
            st.success(f"✅ Se encontraron {len(resultados_ordenados)} registro(s)")
    
    def etiqueta_rrvsac(valor):
        if valor == 'ACTIVO':
            return '<span class="status-active">✅ ACTIVO EN PLATAFORMA</span>'
        else:
            return '<span class="status-inactive">❌ NO ACTIVO EN PLATAFORMA</span>'

    # Mostrar etiqueta de estado solo después de la búsqueda
    if buscar_btn and placa_buscar.strip() and 'rrvsac_status' in locals():
        st.markdown(f'<div style="text-align:center;margin-bottom:18px;">{etiqueta_rrvsac(rrvsac_status)}</div>', unsafe_allow_html=True)

    # Mostrar resultados si existen
    if st.session_state.resultados_actuales:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        col1, col2 = st.columns([2, 1])
        with col1:
            st.subheader("📊 Resultados Encontrados")
        with col2:
            st.metric("Total Registros", len(st.session_state.resultados_actuales))
        df_resultados = pd.DataFrame([
            {
                'FECHA': resultado['fecha'],
                'PLACA': resultado['placa'],
                'EMPRESA': resultado['empresa'],
                'ÚLTIMO ESTADO': resultado['trabajo'],
                'SISTEMA': resultado['sistema'],
                'HOJA': resultado['hoja']
            }
            for resultado in st.session_state.resultados_actuales
        ])
        st.dataframe(
            df_resultados,
            use_container_width=True,
            hide_index=True
        )
        st.subheader("🔍 Detalles Completos")
        for i, resultado in enumerate(st.session_state.resultados_actuales):
            orden_cronologico = "🕒 Más Reciente" if i == 0 else f"📅 Registro #{i+1}"
            with st.expander(f"{orden_cronologico} - Placa: {resultado['placa']} ({resultado['fecha']})"):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**📍 Ubicación del Registro**")
                    st.write(f"**Hoja:** {resultado['hoja']}")
                    st.write(f"**Sistema:** {resultado['sistema']}")
                    st.write(f"**Fila:** {resultado['fila']}")
                with col2:
                    st.markdown("**📊 Información Principal**")
                    st.write(f"**Placa:** {resultado['placa']}")
                    st.write(f"**Fecha:** {resultado['fecha']}")
                    st.write(f"**Empresa:** {resultado['empresa']}")
                    st.write(f"**Estado:** {resultado['trabajo']}")
                st.markdown("**📄 Datos Completos de la Fila**")
                df_detalle = pd.DataFrame({
                    'Campo': resultado['encabezados'],
                    'Valor': resultado['datos_completos']
                })
                st.dataframe(df_detalle, use_container_width=True, hide_index=True)
                excel_bytes = app.crear_excel_bytes(resultado)
                if excel_bytes:
                    st.download_button(
                        label=f"📥 Descargar Excel - Placa {resultado['placa']}",
                        data=excel_bytes,
                        file_name=f"placa_{resultado['placa']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{i}"
                    )
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer
    st.markdown(
        f"<div class='footer'>"
        f"🕒 Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | "
        "🔗 Sistema RRV - Búsqueda de Placas"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
