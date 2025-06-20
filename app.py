import streamlit as st
import gspread
import os
from datetime import datetime
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import io
import json
import tempfile

class BuscadorPlacasWeb:
    def __init__(self):
        self.gc = None
        self.credenciales_path = None
        if 'resultados_actuales' not in st.session_state:
            st.session_state.resultados_actuales = []
        self.detectar_credenciales()
    
    def detectar_credenciales(self):
        """Detecta credenciales desde secrets o archivo local"""
        # Primero intentar desde Streamlit secrets (para producción en Streamlit Cloud)
        try:
            if hasattr(st, 'secrets') and 'gcp_service_account' in st.secrets:
                # Crear archivo temporal con las credenciales
                credentials = dict(st.secrets['gcp_service_account'])
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                    json.dump(credentials, f)
                    self.credenciales_path = f.name
                return
        except Exception:
            pass
        
        # Fallback a archivo local (para desarrollo)
        archivos_json = glob.glob("*.json")
        if archivos_json:
            self.credenciales_path = archivos_json[0]
    
    def conectar_google_sheets(self):
        """Conecta con Google Sheets"""
        try:
            if not self.credenciales_path or not os.path.exists(self.credenciales_path):
                raise FileNotFoundError("Archivo de conexión no encontrado")
            
            self.gc = gspread.service_account(filename=self.credenciales_path)
            return True
        except Exception as e:
            st.error(f"Error de conexión: {str(e)}")
            return False
    
    def buscar_placas_en_drive(self, placa_buscar):
        """Busca una placa en todas las hojas RRV"""
        if not self.gc:
            if not self.conectar_google_sheets():
                return []
        
        try:
            with st.spinner('Buscando en Google Sheets...'):
                todas_las_hojas = self.gc.openall()
                hojas_rrv = [hoja for hoja in todas_las_hojas if "RRV" in hoja.title]
                
                if not hojas_rrv:
                    st.warning("No se encontraron hojas con 'RRV' en el nombre")
                    return []
                
                resultados = []
                progress_bar = st.progress(0)
                
                for idx, hoja in enumerate(hojas_rrv):
                    try:
                        worksheets = hoja.worksheets()
                        
                        for worksheet in worksheets:
                            try:
                                data = worksheet.get_all_values()
                                if not data or len(data) < 2:
                                    continue
                                
                                encabezados = data[0]
                                filas_datos = data[1:]
                                
                                filas_encontradas = self.buscar_placa_en_hoja(
                                    filas_datos, encabezados, placa_buscar, hoja.title, worksheet.title
                                )
                                
                                if filas_encontradas:
                                    resultados.extend(filas_encontradas)
                                    
                            except Exception as e:
                                continue
                                
                    except Exception as e:
                        continue
                    
                    progress_bar.progress((idx + 1) / len(hojas_rrv))
                
                progress_bar.empty()
                return resultados
                
        except Exception as e:
            st.error(f"Error durante la búsqueda: {str(e)}")
            return []
    
    def buscar_placa_en_hoja(self, filas_datos, encabezados, placa_buscar, nombre_spreadsheet, nombre_worksheet):
        """Busca una placa en una hoja específica"""
        resultados = []
        
        # Buscar columnas de placa
        columnas_placa = []
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['placa', 'patente', 'matricula', 'vehiculo', 'numero de vehiculo']):
                columnas_placa.append(i)
        
        if not columnas_placa:
            columnas_placa = list(range(min(3, len(encabezados))))
        
        # Buscar la placa
        for num_fila, fila in enumerate(filas_datos, start=2):
            for col_placa in columnas_placa:
                if col_placa < len(fila):
                    valor_celda = str(fila[col_placa]).strip()
                    if placa_buscar.upper() in valor_celda.upper():
                        # Encontrar columnas específicas
                        fecha_col = self.encontrar_columna_fecha(encabezados)
                        proyecto_col = self.encontrar_columna_proyecto(encabezados)
                        empresa_col = self.encontrar_columna_empresa(encabezados)
                        sistema_col = self.encontrar_columna_sistema(encabezados)
                        trabajo_col = self.encontrar_columna_trabajo(encabezados)
                        
                        resultado = {
                            'hoja': nombre_spreadsheet,
                            'pestana': nombre_worksheet,
                            'fila': num_fila,
                            'placa': valor_celda,
                            'fecha': fila[fecha_col] if fecha_col < len(fila) else "No disponible",
                            'proyecto': fila[proyecto_col] if proyecto_col < len(fila) else "No disponible",
                            'empresa': fila[empresa_col] if empresa_col < len(fila) else "No disponible",
                            'sistema': fila[sistema_col] if sistema_col < len(fila) else "No disponible",
                            'trabajo': fila[trabajo_col] if trabajo_col < len(fila) else "No disponible",
                            'datos_completos': fila,
                            'encabezados': encabezados
                        }
                        resultados.append(resultado)
                        break
        
        return resultados
    
    def encontrar_columna_fecha(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['fecha', 'date', 'dia', 'hora', 'fecha de ingreso']):
                return i
        return 1 if len(encabezados) > 1 else 0
    
    def encontrar_columna_proyecto(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if 'proyecto' in encabezado_lower:
                return i
        return 2 if len(encabezados) > 2 else 0
    
    def encontrar_columna_empresa(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['empresa', 'nombre', 'cliente']):
                return i
        return 3 if len(encabezados) > 3 else 0
    
    def encontrar_columna_sistema(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if 'sistema' in encabezado_lower:
                return i
        return 4 if len(encabezados) > 4 else 0
    
    def encontrar_columna_trabajo(self, encabezados):
        for i, encabezado in enumerate(encabezados):
            encabezado_lower = str(encabezado).lower()
            if any(palabra in encabezado_lower for palabra in ['tipo de trabajo', 'estado', 'status', 'situacion', 'condicion']):
                return i
        return 5 if len(encabezados) > 5 else 0
    
    def crear_excel_bytes(self, resultado):
        """Crea un archivo Excel en memoria y devuelve los bytes"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Placa {resultado['placa']}"
            
            # Estilos
            titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            subtitulo_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=10)
            
            titulo_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
            subtitulo_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws['A1'] = f"INFORMACIÓN DE LA PLACA: {resultado['placa']}"
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws.merge_cells('A1:C1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Información de ubicación
            ws['A3'] = "UBICACIÓN DEL REGISTRO"
            ws['A3'].font = subtitulo_font
            ws['A3'].fill = subtitulo_fill
            ws['A3'].border = border
            
            ws['A4'] = "Hoja:"
            ws['B4'] = resultado['hoja']
            ws['A5'] = "Pestaña:"
            ws['B5'] = resultado['pestana']
            ws['A6'] = "Fila:"
            ws['B6'] = resultado['fila']
            
            # Aplicar estilos a la información de ubicación
            for row in range(4, 7):
                ws[f'A{row}'].font = normal_font
                ws[f'A{row}'].border = border
                ws[f'B{row}'].font = normal_font
                ws[f'B{row}'].border = border
            
            # Datos completos de la fila
            ws['A8'] = "DATOS COMPLETOS DE LA FILA"
            ws['A8'].font = subtitulo_font
            ws['A8'].fill = subtitulo_fill
            ws['A8'].border = border
            ws.merge_cells('A8:C8')
            
            # Encabezados de la tabla de datos
            ws['A9'] = "Campo"
            ws['B9'] = "Valor"
            ws['A9'].font = subtitulo_font
            ws['B9'].font = subtitulo_font
            ws['A9'].fill = subtitulo_fill
            ws['B9'].fill = subtitulo_fill
            ws['A9'].border = border
            ws['B9'].border = border
            
            # Insertar todos los datos de la fila
            row_num = 10
            for encabezado, valor in zip(resultado['encabezados'], resultado['datos_completos']):
                ws[f'A{row_num}'] = encabezado
                ws[f'B{row_num}'] = valor
                ws[f'A{row_num}'].font = normal_font
                ws[f'B{row_num}'].font = normal_font
                ws[f'A{row_num}'].border = border
                ws[f'B{row_num}'].border = border
                row_num += 1
            
            # Información del archivo
            ws[f'A{row_num + 1}'] = f"Archivo generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws[f'A{row_num + 1}'].font = Font(name='Arial', size=9, italic=True)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            wb.close()
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error al crear archivo Excel: {str(e)}")
            return None

def main():
    # Configuración de la página
    st.set_page_config(
        page_title="BUSCADOR RRV",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # CSS personalizado para mejorar el diseño
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #2196F3 0%, #1976D2 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
    }
    .search-container {
        background: white;
        padding: 2rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
    }
    .results-container {
        background: white;
        padding: 2rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header principal
    st.markdown("""
    <div class="main-header">
        <h1>🔍 BUSCADOR RRV</h1>
        <p>Sistema de búsqueda de placas en Google Sheets</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Inicializar la aplicación
    app = BuscadorPlacasWeb()
    
    # Verificar credenciales
    if not app.credenciales_path:
        st.error("❌ No se encontraron credenciales. Contacta al administrador para configurar el acceso.")
        st.info("💡 Para desarrolladores: Configura las credenciales en Streamlit Cloud Secrets o agrega un archivo JSON local.")
        return
    
    st.success("✅ Sistema conectado y listo para buscar.")
    
    # Container de búsqueda
    with st.container():
        st.markdown('<div class="search-container">', unsafe_allow_html=True)
        
        st.subheader("📋 Buscar Placa")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            placa_buscar = st.text_input(
                "Ingresa la placa a buscar:",
                placeholder="Ej: ABC-123",
                key="placa_input"
            )
        
        with col2:
            st.write("")  # Espaciado
            buscar_btn = st.button("🔍 Buscar", type="primary", use_container_width=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Ejecutar búsqueda
    if buscar_btn and placa_buscar.strip():
        resultados = app.buscar_placas_en_drive(placa_buscar.strip())
        st.session_state.resultados_actuales = resultados
        
        if not resultados:
            st.warning("⚠️ No se encontró esta placa en el sistema")
        else:
            st.success(f"✅ Se encontraron {len(resultados)} registro(s)")
    
    # Mostrar resultados si existen
    if st.session_state.resultados_actuales:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        
        # Header de resultados
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.subheader("📊 Resultados Encontrados")
        
        with col2:
            st.metric("Total Registros", len(st.session_state.resultados_actuales))
        
        # Crear DataFrame para mostrar los resultados
        df_resultados = pd.DataFrame([
            {
                'FECHA': resultado['fecha'],
                'PLACA': resultado['placa'],
                'EMPRESA': resultado['empresa'],
                'ÚLTIMO ESTADO': resultado['trabajo'],
                'SERVICIO': resultado['pestana'],
                'HOJA': resultado['hoja']
            }
            for resultado in st.session_state.resultados_actuales
        ])
        
        # Mostrar tabla
        st.dataframe(
            df_resultados,
            use_container_width=True,
            hide_index=True
        )
        
        # Mostrar detalles expandibles
        st.subheader("🔍 Detalles Completos")
        
        for i, resultado in enumerate(st.session_state.resultados_actuales):
            with st.expander(f"📋 Registro #{i+1} - Placa: {resultado['placa']} ({resultado['fecha']})"):
                
                # Información básica
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**📍 Ubicación del Registro**")
                    st.write(f"**Hoja:** {resultado['hoja']}")
                    st.write(f"**Pestaña:** {resultado['pestana']}")
                    st.write(f"**Fila:** {resultado['fila']}")
                
                with col2:
                    st.markdown("**📊 Información Principal**")
                    st.write(f"**Placa:** {resultado['placa']}")
                    st.write(f"**Fecha:** {resultado['fecha']}")
                    st.write(f"**Empresa:** {resultado['empresa']}")
                    st.write(f"**Estado:** {resultado['trabajo']}")
                
                # Todos los datos de la fila
                st.markdown("**📄 Datos Completos de la Fila**")
                
                # Crear DataFrame con todos los datos
                df_detalle = pd.DataFrame({
                    'Campo': resultado['encabezados'],
                    'Valor': resultado['datos_completos']
                })
                
                st.dataframe(df_detalle, use_container_width=True, hide_index=True)
                
                # Botón de descarga individual
                excel_bytes = app.crear_excel_bytes(resultado)
                if excel_bytes:
                    st.download_button(
                        label=f"📥 Descargar Excel - Placa {resultado['placa']}",
                        data=excel_bytes,
                        file_name=f"placa_{resultado['placa']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{i}"
                    )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666; font-size: 0.9em;'>"
        f"🕒 Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | "
        "🔗 Ejecutándose en GitHub Codespaces"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
